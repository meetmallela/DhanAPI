"""
generate_assessment_v3.py
--------------------------
Final assessment Excel -- v3.
Incorporates all changes from both strategy assessment documents
(StrategyAssesment_31May2026_11_51.md  and  StrategyAssesment_31May2026_12_36.md).

Output:
    MasterConfiguration/reports/Strategy_Assessment_v3.xlsx

Sheets:
  1. Executive_Summary   -- 3-way comparison: Original (v1) -> Proposed (v2) -> Final (v3)
  2. Final_Configs       -- STRATEGY_CONFIGS with final parameters + rationale
  3. Strategy_Evolution  -- Full metrics table v1->v2->v3 for every pattern x symbol
  4. GammaSqueeze        -- New strategy deep-dive: day-of-week, all signals, risk/reward
  5. GapFill_Deep        -- BANKNIFTY + SENSEX gap analysis (immediate open, v1 entry)
  6. NIFTY_6M            -- Daily log with final v3 signals (training window)
  7. NIFTY_7to12M        -- NIFTY test window
  8. BANKNIFTY_6M        -- BANKNIFTY training
  9. BANKNIFTY_7to12M    -- BANKNIFTY test
 10. SENSEX_6M           -- SENSEX training
 11. SENSEX_7to12M       -- SENSEX test
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_ROOT   = Path(__file__).parent
_MASTER = _ROOT.parent / "MasterConfiguration"
_LIB    = _MASTER / "lib"
if str(_LIB) not in sys.path:
    sys.path.insert(0, str(_LIB))

from historical_data_fetch import load
from pattern_discovery import (
    add_indicators, summarise,
    find_expiry_blast, find_orb15, find_gap_fill,
    find_atr_squeeze, find_vwap_reclaim,
)
from pattern_discovery_v2 import (
    add_indicators_v2,
    find_expiry_blast_v2, find_orb15_v2, find_gap_fill_v2,
    find_atr_squeeze_v2, find_vwap_reclaim_v2,
    find_gamma_squeeze,
)

OUTPUT = _MASTER / "reports" / "Strategy_Assessment_v3.xlsx"
# (path assembled from Path objects -- no raw backslash strings needed)
OUTPUT.parent.mkdir(parents=True, exist_ok=True)

TODAY      = datetime.now()
TRAIN_TO   = TODAY.strftime("%Y-%m-%d")
TRAIN_FROM = (TODAY - timedelta(days=6 * 31)).strftime("%Y-%m-%d")
TEST_TO    = TRAIN_FROM
TEST_FROM  = (TODAY - timedelta(days=12 * 31)).strftime("%Y-%m-%d")

SYMBOLS = ["NIFTY", "BANKNIFTY", "SENSEX"]

# ── Colour tokens ─────────────────────────────────────────────────────────────
def _fill(hex_colour): return PatternFill("solid", fgColor=hex_colour)

C_TITLE      = _fill("1F4E79")
C_HDR_BLUE   = _fill("2E75B6")
C_HDR_DARK   = _fill("1A3A5C")
C_HDR_SEC    = _fill("4472C4")
C_GREEN_DARK = _fill("1E7B34")
C_GREEN_MED  = _fill("70AD47")
C_GREEN_LT   = _fill("C6EFCE")
C_RED_LT     = _fill("FFC7CE")
C_RED_MED    = _fill("C00000")
C_YELLOW     = _fill("FFEB9C")
C_BLUE_LT    = _fill("BDD7EE")
C_GREY_LT    = _fill("F2F2F2")
C_ORANGE_LT  = _fill("FCE4D6")
C_PURPLE_LT  = _fill("E2EFDA")
C_WHITE      = _fill("FFFFFF")

F_WHITE_BOLD = Font(color="FFFFFF", bold=True)
F_BOLD       = Font(bold=True)
F_TITLE      = Font(bold=True, size=14, color="FFFFFF")
F_SUBTITLE   = Font(bold=True, size=11, color="FFFFFF")
F_HDR        = Font(bold=True, color="FFFFFF")

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

PATTERN_COLOURS = {
    "ExpiryBlast": "FF7043", "ORB15":  "29B6F6",
    "GapFill":     "AB47BC", "ATRSqueeze": "26A69A",
    "VWAPReclaim": "FFA726", "GammaSqueeze": "E91E63",
}

# ── Cell helpers ──────────────────────────────────────────────────────────────

def hdr(ws, r, c, val, bg="2E75B6", fg="FFFFFF", bold=True,
        size=10, halign="center", wrap=True, rowh=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill      = _fill(bg)
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=halign, vertical="center",
                                wrap_text=wrap)
    cell.border    = THIN
    if rowh:
        ws.row_dimensions[r].height = rowh
    return cell


def cell(ws, r, c, val, fill=None, bold=False, halign="center",
         fmt=None, font_colour="000000"):
    ce = ws.cell(row=r, column=c, value=val)
    if fill:
        ce.fill = fill
    ce.font      = Font(bold=bold, color=font_colour)
    ce.alignment = Alignment(horizontal=halign, vertical="center")
    ce.border    = THIN
    if fmt:
        ce.number_format = fmt
    return ce


def title_row(ws, r, cols_span, text, bg="1F4E79", size=14, height=30):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols_span)
    c = ws.cell(row=r, column=1, value=text)
    c.fill = _fill(bg)
    c.font = Font(bold=True, size=size, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = height
    return c


def section_title(ws, r, cols_span, text, bg="4472C4"):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols_span)
    c = ws.cell(row=r, column=1, value=text)
    c.fill = _fill(bg)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 20
    return c


def freeze(ws, r=2, c=1):
    ws.freeze_panes = ws.cell(row=r, column=c)


def col_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def delta_fill(new, old):
    if new is None or old is None:
        return C_GREY_LT
    if new > old * 1.03:
        return C_GREEN_LT
    if new < old * 0.97:
        return C_RED_LT
    return C_YELLOW


def wr_fill(wr):
    if wr >= 60: return C_GREEN_MED
    if wr >= 50: return C_GREEN_LT
    if wr >= 45: return C_YELLOW
    return C_RED_LT


def exp_fill(e):
    if e > 50:  return C_GREEN_MED
    if e > 10:  return C_GREEN_LT
    if e > 0:   return C_YELLOW
    return C_RED_LT


def grade(wr, exp, n):
    if n < 5:               return "INSUFFICIENT", C_GREY_LT
    if exp > 50 and wr >= 60: return "STRONG EDGE",  C_GREEN_MED
    if exp > 10 and wr >= 50: return "EDGE",          C_GREEN_LT
    if exp > 0:               return "MARGINAL",      C_YELLOW
    return "NO EDGE", C_RED_LT


# ── Data loading ──────────────────────────────────────────────────────────────

def load_and_compute(symbol, from_d, to_d):
    df = load(symbol, "5min", from_d, to_d)
    if df.empty:
        return pd.DataFrame()
    df = df.sort_values("timestamp").reset_index(drop=True)
    return add_indicators_v2(df)


def run_v1(df, symbol):
    if df.empty:
        return {}
    return {
        "ExpiryBlast": find_expiry_blast(df, symbol),
        "ORB15":       find_orb15(df, symbol),
        "GapFill":     find_gap_fill(df, symbol),
        "ATRSqueeze":  find_atr_squeeze(df, symbol),
        "VWAPReclaim": find_vwap_reclaim(df, symbol),
    }


def run_v3(df, symbol):
    if df.empty:
        return {}
    active = lambda d: d[d.get("skipped", pd.Series([False]*len(d))) == False] if "skipped" in d.columns else d
    return {
        "ExpiryBlast_v2": find_expiry_blast_v2(df, symbol),
        "ORB15_v2":       find_orb15_v2(df, symbol).pipe(active),
        "GapFill_v2":     find_gap_fill_v2(df, symbol),
        "ATRSqueeze_v2":  find_atr_squeeze_v2(df, symbol).pipe(active),
        "VWAPReclaim_v2": find_vwap_reclaim_v2(df, symbol),
        "GammaSqueeze":   find_gamma_squeeze(df, symbol),
    }


V3_ENABLED = {
    "ExpiryBlast_v2": ["BANKNIFTY", "SENSEX"],
    "ORB15_v2":       ["NIFTY", "BANKNIFTY"],
    "GapFill_v2":     ["BANKNIFTY", "SENSEX"],
    "ATRSqueeze_v2":  ["NIFTY", "BANKNIFTY", "SENSEX"],
    "VWAPReclaim_v2": ["NIFTY", "BANKNIFTY", "SENSEX"],
    "GammaSqueeze":   ["NIFTY", "BANKNIFTY", "SENSEX"],
}

V1_PATTERNS = ["ExpiryBlast", "ORB15", "GapFill", "ATRSqueeze", "VWAPReclaim"]
V3_PATTERNS = ["ExpiryBlast_v2", "ORB15_v2", "GapFill_v2",
                "ATRSqueeze_v2", "VWAPReclaim_v2", "GammaSqueeze"]


# ── Sheet 1: Executive Summary ────────────────────────────────────────────────

def write_executive_summary(ws, all_data):
    ws.title = "Executive_Summary"
    ws.sheet_view.showGridLines = False

    title_row(ws, 1, 16,
              "Strategy Assessment v3  |  Final Configuration  |  "
              "NIFTY + BANKNIFTY + SENSEX  |  May 31 2026", size=13, height=32)

    # Key stats block
    r = 3
    section_title(ws, r, 16, "A.  What Changed -- v1 -> v2 -> v3")
    r += 1
    changes = [
        ("ExpiryBlast", "Retired on NIFTY; BANKNIFTY+SENSEX use expanded 14:45-15:15 window + BB-coiling pre-filter (<15th pctile). Rare but high-conviction when it fires."),
        ("ORB15",       "GAP FILTER added: skip days with opening gap >0.5%. SENSEX disabled. Result: NIFTY 49%->77% WR, BANKNIFTY 48%->75% WR (out-of-sample)."),
        ("GapFill",     "NIFTY disabled (low fill rate). BANKNIFTY+SENSEX: v2 5-min-reversal entry REVERTED to v1 immediate-open entry -- 5-min filter delayed entry past the fill, collapsing test expectancy from +81 to -35 pts on BANKNIFTY."),
        ("ATRSqueeze",  "Volume 1.8x SMA confirm added (bypassed for zero-vol spot indices). VIX-adaptive percentile parameter ready for live (15th if VIX<15, 25th if VIX>15). Unchanged expectancy -- all spot indices bypass vol filter."),
        ("VWAPReclaim", "Confirm-bar threshold raised: 3 bars (NIFTY) / 5 bars (BANKNIFTY, SENSEX). Reclaim candle body% > 65% required. Fewer but cleaner signals."),
        ("GammaSqueeze","NEW strategy. Mon/Wed/Fri 14:45-15:15 IST. ADX>45 + RSI 35-50 + 3 consecutive bars. BANKNIFTY test: 100% WR, +437 pts avg. SENSEX test: 57% WR, +86 pts."),
    ]
    hdrs = ["Strategy", "Change Summary"]
    for ci, h in enumerate(hdrs, 1):
        hdr(ws, r, ci, h)
    col_w(ws, [18, 110])
    r += 1
    for pat, txt in changes:
        bg = _fill(PATTERN_COLOURS.get(pat.split("_")[0] if "_" in pat else pat, "DDDDDD") + "33")
        cell(ws, r, 1, pat,   fill=bg, bold=True, halign="left")
        cell(ws, r, 2, txt,   fill=None, halign="left")
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    section_title(ws, r, 16, "B.  Final Deployment Decision Matrix")
    r += 1
    col_hdrs = ["Strategy", "NIFTY", "BANKNIFTY", "SENSEX",
                "Train WR (best)", "Test WR (best)", "Test Exp (best)", "Action"]
    for ci, h in enumerate(col_hdrs, 1):
        hdr(ws, r, ci, h)
    ws.row_dimensions[r].height = 28
    r += 1

    deploy_matrix = [
        ("ExpiryBlast_v2", "DISABLED", "ACTIVE",    "ACTIVE",
         "100% (SENSEX)", "100% (SENSEX)", "+156 pts", "Monitor -- too few signals for stat confidence"),
        ("ORB15_v2",       "ACTIVE",   "ACTIVE",    "DISABLED",
         "66% (NIFTY)",   "77% (NIFTY)",  "+55 pts (BN)", "DEPLOY -- strong consistent edge"),
        ("GapFill_v2",     "DISABLED", "ACTIVE",    "ACTIVE",
         "58% (BN)",      "67% (BN)",     "+81 pts (BN)", "DEPLOY -- v1 entry retained"),
        ("ATRSqueeze_v2",  "ACTIVE",   "ACTIVE",    "ACTIVE",
         "54% (NIFTY)",   "55% (NIFTY)",  "+79 pts (SX)", "DEPLOY -- workhorse, N~500 per symbol"),
        ("VWAPReclaim_v2", "ACTIVE",   "ACTIVE",    "ACTIVE",
         "52% (SENSEX)",  "51% (SENSEX)", "+59 pts (SX)", "DEPLOY -- improved with 5-bar filter"),
        ("GammaSqueeze",   "WATCH",    "WATCH",     "WATCH",
         "100% (BN)",     "100% (BN)",    "+437 pts (BN)","WATCH -- spectacular but N<10; paper-trade first"),
    ]
    action_colours = {
        "DEPLOY":  C_GREEN_LT, "Monitor": C_YELLOW,
        "WATCH":   C_BLUE_LT,  "DISABLED": C_GREY_LT,
    }
    for strat, nifty, bnk, sx, trWR, teWR, teExp, action in deploy_matrix:
        def _status_fill(s):
            if s == "ACTIVE":    return C_GREEN_LT
            if s == "DISABLED":  return C_GREY_LT
            if s == "WATCH":     return C_BLUE_LT
            return C_YELLOW
        cell(ws, r, 1, strat,  bold=True, halign="left")
        cell(ws, r, 2, nifty,  fill=_status_fill(nifty))
        cell(ws, r, 3, bnk,    fill=_status_fill(bnk))
        cell(ws, r, 4, sx,     fill=_status_fill(sx))
        cell(ws, r, 5, trWR,   fill=C_GREEN_LT)
        cell(ws, r, 6, teWR,   fill=C_GREEN_MED)
        cell(ws, r, 7, teExp,  fill=C_GREEN_MED)
        ac_fill = C_GREEN_LT if "DEPLOY" in action else C_BLUE_LT if "WATCH" in action else C_YELLOW
        cell(ws, r, 8, action, fill=ac_fill, halign="left")
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    section_title(ws, r, 16, "C.  Critical Finding -- GapFill v2 Reversion")
    r += 1
    warning = (
        "The 5-minute reversal-candle entry filter introduced in v2 was REVERTED in v3 for both BANKNIFTY and SENSEX. "
        "Waiting for the first 5m candle to confirm the gap-fill direction delayed entry until most of the fill had already occurred. "
        "This led to entering right before secondary whipsaws, collapsing test expectancy: "
        "BANKNIFTY: +81.3 pts (v1) -> -35.1 pts (v2). SENSEX: +48.5 pts (v1) -> -43.9 pts (v2). "
        "v3 uses immediate open entry (v1 behaviour) restricted to BANKNIFTY and SENSEX only."
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16)
    c = ws.cell(row=r, column=1, value=warning)
    c.fill      = _fill("FFF2CC")
    c.font      = Font(bold=False, size=10, color="7F6000")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = THIN
    ws.row_dimensions[r].height = 60

    freeze(ws, 2, 1)


# ── Sheet 2: Final Configs ────────────────────────────────────────────────────

def write_final_configs(ws):
    ws.title = "Final_Configs"
    ws.sheet_view.showGridLines = False

    title_row(ws, 1, 8,
              "Final Strategy Configurations -- v3  (STRATEGY_CONFIGS as of 2026-05-31)", height=28)

    configs = [
        # Pattern, Index, Enabled, Trigger/Window, Key Parameters, Entry, SL/TP, Notes
        ("ExpiryBlast_v2", "NIFTY",     "DISABLED", "N/A",
         "N/A",
         "N/A", "N/A",
         "Insufficient signals in current regime -- BB coiling filter too strict"),

        ("ExpiryBlast_v2", "BANKNIFTY", "ACTIVE",   "14:45-15:15 IST",
         "BB width < 15th pctile (50-bar) | move > 1.5x ATR14 | body% > 0.60 | RSI 30-70",
         "Close of qualifying bar", "SL: 1x ATR below entry | Target: next 6 bars max",
         "Expiry day only (Wednesday for BANKNIFTY)"),

        ("ExpiryBlast_v2", "SENSEX",    "ACTIVE",   "14:45-15:15 IST",
         "BB width < 15th pctile (50-bar) | move > 1.5x ATR14 | body% > 0.60 | RSI 30-70",
         "Close of qualifying bar", "SL: 1x ATR below entry | Target: next 6 bars max",
         "Expiry day only (Tuesday for SENSEX)"),

        ("ORB15_v2", "NIFTY",     "ACTIVE", "9:15-9:29 ORB | Entry from 9:30+",
         "Gap filter: opening gap < 0.5% | EMA9/21 confirm | RSI 30-70",
         "Close of breakout bar", "SL: opposite ORB level | Win: >1 ATR in 15 min OR >0.5 ATR in 30 min",
         "Max opening gap 0.5%. Sluggish trades cut at 15 min."),

        ("ORB15_v2", "BANKNIFTY", "ACTIVE", "9:15-9:29 ORB | Entry from 9:30+",
         "Gap filter: opening gap < 0.5% | EMA9/21 confirm | RSI 30-70",
         "Close of breakout bar", "SL: opposite ORB level | Win: >1 ATR in 15 min OR >0.5 ATR in 30 min",
         "Higher ATR moves make this the best ORB vehicle"),

        ("ORB15_v2", "SENSEX",    "DISABLED", "N/A",
         "N/A", "N/A", "N/A",
         "No edge in testing (43% WR). Capital redirected to GapFill and ATRSqueeze."),

        ("GapFill_v2", "NIFTY",     "DISABLED", "N/A",
         "N/A", "N/A", "N/A",
         "Low fill rate -- NIFTY diversification suppresses rapid gap fills"),

        ("GapFill_v2", "BANKNIFTY", "ACTIVE", "9:15 open",
         "Min gap: 0.3% vs prev close | Entry: IMMEDIATE at open (v1 -- 5-min filter reverted)",
         "Market open price", "TP: prev close | SL: 0.3% beyond open | Time exit: 90 min",
         "Test WR: 66.7%, +81.3 pts. Do NOT re-add 5-min filter."),

        ("GapFill_v2", "SENSEX",    "ACTIVE", "9:15 open",
         "Min gap: 0.3% vs prev close | Entry: IMMEDIATE at open (v1 -- 5-min filter reverted)",
         "Market open price", "TP: prev close | SL: 0.3% beyond open | Time exit: 90 min",
         "Test WR: 58.1%, +48.5 pts. Same reversion logic."),

        ("ATRSqueeze_v2", "NIFTY",     "ACTIVE", "Any time 9:30-14:45",
         "BB width < 20th pctile (50-bar, VIX-adaptive*) | Close crosses BB-mid | RSI not exhausted (<72/>28)",
         "Close of breakout bar", "SL: 1.5x ATR | Target: next 12 bars max favourable",
         "*Live: use 15th pctile if VIX<15, 25th if VIX>15. Volume bypass (spot index vol=0)."),

        ("ATRSqueeze_v2", "BANKNIFTY", "ACTIVE", "Any time 9:30-14:45",
         "BB width < 20th pctile (VIX-adaptive*) | BB-mid cross | RSI not exhausted",
         "Close of breakout bar", "SL: 1.5x ATR | Target: next 12 bars",
         "Test: +61.2 pts expectancy. Highest trade count."),

        ("ATRSqueeze_v2", "SENSEX",    "ACTIVE", "Any time 9:30-14:45",
         "BB width < 20th pctile (VIX-adaptive*) | BB-mid cross | RSI not exhausted",
         "Close of breakout bar", "SL: 1.5x ATR | Target: next 12 bars",
         "Test: +78.5 pts. Best ATRSqueeze expectancy of all 3 indices."),

        ("VWAPReclaim_v2", "NIFTY",     "ACTIVE", "9:30-14:30",
         "3 consecutive bars on wrong side of VWAP | Cross with RSI>45 (buy) or <55 (sell) | Body%>65%",
         "Close of reclaim bar", "SL: 1x ATR | Target: next 6 bars",
         "3-bar confirm (lower threshold for less volatile NIFTY)"),

        ("VWAPReclaim_v2", "BANKNIFTY", "ACTIVE", "9:30-14:30",
         "5 consecutive bars on wrong side of VWAP | Cross with RSI>45/<55 | Body%>65%",
         "Close of reclaim bar", "SL: 1x ATR | Target: next 6 bars",
         "5-bar confirm prevents high-beta stop-hunt false reclaims"),

        ("VWAPReclaim_v2", "SENSEX",    "ACTIVE", "9:30-14:30",
         "5 consecutive bars on wrong side of VWAP | Cross with RSI>45/<55 | Body%>65%",
         "Close of reclaim bar", "SL: 1x ATR | Target: next 6 bars",
         "Test: +59.4 pts expectancy after 5-bar upgrade"),

        ("GammaSqueeze",   "NIFTY",     "WATCH",  "14:45-15:15, Mon/Wed/Fri only",
         "ADX14>45 | RSI 35-50 | 3 consecutive bars same direction | Vol surge (bypass spot)",
         "Close of qualifying bar", "Strict trailing SL | Hard exit by 15:20",
         "Test: 67% WR, +26 pts. Small sample (N=6) -- paper-trade first."),

        ("GammaSqueeze",   "BANKNIFTY", "WATCH",  "14:45-15:15, Mon/Wed/Fri only",
         "ADX14>45 | RSI 35-50 | 3 consecutive bars | Vol surge (bypass spot)",
         "Close of qualifying bar", "Strict trailing SL | Hard exit by 15:20",
         "Test: 100% WR, +437 pts. Spectacular but N=2. Paper-trade first."),

        ("GammaSqueeze",   "SENSEX",    "WATCH",  "14:45-15:15, Mon/Wed/Fri only",
         "ADX14>45 | RSI 35-50 | 3 consecutive bars | Vol surge (bypass spot)",
         "Close of qualifying bar", "Strict trailing SL | Hard exit by 15:20",
         "Test: 57% WR, +86 pts, N=7. Best sample size -- promote to ACTIVE after 20 signals."),
    ]

    col_hdrs = ["Strategy", "Index", "Status", "Time Window",
                "Conditions", "Entry", "SL / TP", "Notes"]
    r = 2
    for ci, h in enumerate(col_hdrs, 1):
        hdr(ws, r, ci, h, rowh=28)
    r += 1

    for row_data in configs:
        strat, idx, status, window, cond, entry, sltp, notes = row_data
        status_fill = (C_GREEN_LT  if status == "ACTIVE"   else
                       C_BLUE_LT   if status == "WATCH"    else
                       C_GREY_LT)
        pat_bg = _fill(PATTERN_COLOURS.get(
            strat.replace("_v2", "").replace("GammaSqueeze", "GammaSqueeze"), "DDDDDD") + "22")
        cell(ws, r, 1, strat,  fill=pat_bg, bold=True,  halign="left")
        cell(ws, r, 2, idx,    fill=pat_bg, bold=True,  halign="center")
        cell(ws, r, 3, status, fill=status_fill, bold=True)
        cell(ws, r, 4, window, halign="left")
        cell(ws, r, 5, cond,   halign="left")
        cell(ws, r, 6, entry,  halign="left")
        cell(ws, r, 7, sltp,   halign="left")
        cell(ws, r, 8, notes,  halign="left",
             fill=C_YELLOW if "REVERTED" in notes or "Do NOT" in notes else None)
        ws.row_dimensions[r].height = 36
        r += 1

    col_w(ws, [18, 11, 10, 20, 60, 26, 38, 55])
    freeze(ws, 3, 1)


# ── Sheet 3: Strategy Evolution (v1->v2->v3) ──────────────────────────────────

def write_evolution(ws, stats):
    """stats: {sym: {window: {v1: {pat: summary}, v3: {pat: summary}}}}"""
    ws.title = "Strategy_Evolution"
    ws.sheet_view.showGridLines = False

    title_row(ws, 1, 20,
              "Strategy Performance Evolution -- Original (v1) -> Proposed (v2) -> Final (v3)", height=28)

    r = 2
    hdrs = [
        "Symbol", "Pattern", "Window",
        "v1 N", "v1 WR%", "v1 Exp",
        "v3 N", "v3 WR%", "v3 Exp",
        "WR Delta", "Exp Delta", "Status", "Grade", "Recommendation",
    ]
    for ci, h in enumerate(hdrs, 1):
        hdr(ws, r, ci, h, rowh=28)
    r += 1

    v1_map = {"ExpiryBlast_v2": "ExpiryBlast", "ORB15_v2": "ORB15",
              "GapFill_v2": "GapFill", "ATRSqueeze_v2": "ATRSqueeze",
              "VWAPReclaim_v2": "VWAPReclaim", "GammaSqueeze": None}

    for sym in SYMBOLS:
        for window in ("Train", "Test"):
            for v3_pat in V3_PATTERNS:
                v1_pat = v1_map.get(v3_pat)
                enabled = sym in V3_ENABLED.get(v3_pat, [])

                v1s = stats.get(sym, {}).get(window, {}).get("v1", {}).get(v1_pat, {}) if v1_pat else {}
                v3s = stats.get(sym, {}).get(window, {}).get("v3", {}).get(v3_pat, {})

                v1_wr  = v1s.get("win_rate", 0)
                v1_exp = v1s.get("expectancy_pts", 0)
                v1_n   = v1s.get("count", 0)
                v3_wr  = v3s.get("win_rate", 0)
                v3_exp = v3s.get("expectancy_pts", 0)
                v3_n   = v3s.get("count", 0)

                wr_d   = round(v3_wr  - v1_wr,  1) if v1_wr  else None
                exp_d  = round(v3_exp - v1_exp, 1) if v1_exp else None

                status = "ACTIVE" if enabled else "DISABLED"
                grd, grd_fill = grade(v3_wr, v3_exp, v3_n)

                rec = ("DEPLOY" if grd in ("STRONG EDGE", "EDGE") and enabled else
                       "WATCH"  if grd == "STRONG EDGE" and not enabled else
                       "DISABLED" if not enabled else
                       "MARGINAL -- monitor")

                row_vals = [
                    sym, v3_pat, window,
                    v1_n or "-", v1_wr or "-", v1_exp or "-",
                    v3_n or "-", v3_wr or "-", v3_exp or "-",
                    wr_d, exp_d, status, grd, rec,
                ]
                fills = [
                    None, None, None,
                    None, wr_fill(v1_wr) if v1_wr else C_GREY_LT,
                    exp_fill(v1_exp) if v1_exp else C_GREY_LT,
                    None, wr_fill(v3_wr) if v3_wr else (C_GREY_LT if not enabled else None),
                    exp_fill(v3_exp) if v3_exp else (C_GREY_LT if not enabled else None),
                    delta_fill(v3_wr, v1_wr) if v1_wr else C_GREY_LT,
                    delta_fill(v3_exp, v1_exp) if v1_exp else C_GREY_LT,
                    C_GREEN_LT if status == "ACTIVE" else C_GREY_LT,
                    grd_fill,
                    C_GREEN_MED if rec == "DEPLOY" else C_BLUE_LT if rec == "WATCH" else C_GREY_LT,
                ]
                for ci, (v, f) in enumerate(zip(row_vals, fills), 1):
                    cell(ws, r, ci, v, fill=f,
                         halign="left" if ci in (1, 2, 3, 12, 13, 14) else "center")
                ws.row_dimensions[r].height = 18
                r += 1

    col_w(ws, [12, 17, 7, 6, 8, 9, 6, 8, 9, 9, 9, 10, 16, 22])
    freeze(ws, 3, 1)


# ── Sheet 4: GammaSqueeze Deep-Dive ───────────────────────────────────────────

def write_gamma_squeeze(ws, gamma_data):
    """gamma_data: {sym: {window: df_signals}}"""
    ws.title = "GammaSqueeze"
    ws.sheet_style = None
    ws.sheet_view.showGridLines = False

    title_row(ws, 1, 14,
              "GammaSqueeze -- New Strategy Deep Dive  |  Mon/Wed/Fri  |  14:45-15:15 IST  |  ADX>45  |  RSI 35-50",
              height=28)

    r = 3
    section_title(ws, r, 14, "A.  Overall Stats by Symbol and Window")
    r += 1
    hdr_vals = ["Symbol", "Window", "N Signals", "Win Rate%", "Avg Win(pts)",
                "Avg Loss(pts)", "Expectancy(pts)", "Best(pts)", "Worst(pts)", "Grade"]
    for ci, h in enumerate(hdr_vals, 1):
        hdr(ws, r, ci, h)
    r += 1

    for sym in SYMBOLS:
        for window in ("Train", "Test"):
            df = gamma_data.get(sym, {}).get(window, pd.DataFrame())
            if df.empty:
                s = {"count": 0, "win_rate": 0, "expectancy_pts": 0,
                     "avg_win_pts": 0, "avg_loss_pts": 0, "best_pts": 0, "worst_pts": 0}
            else:
                s = summarise(df)
                s["best_pts"]  = round(df["outcome_pts"].max(), 1) if not df.empty else 0
                s["worst_pts"] = round(df["outcome_pts"].min(), 1) if not df.empty else 0

            grd, grd_fill = grade(s["win_rate"], s["expectancy_pts"], s["count"])
            vals = [sym, window, s["count"], s["win_rate"], s["avg_win_pts"],
                    s["avg_loss_pts"], s["expectancy_pts"], s.get("best_pts", 0),
                    s.get("worst_pts", 0), grd]
            for ci, v in enumerate(vals, 1):
                f = (exp_fill(s["expectancy_pts"]) if ci == 7 else
                     wr_fill(s["win_rate"]) if ci == 4 else
                     grd_fill if ci == 10 else None)
                cell(ws, r, ci, v, fill=f)
            ws.row_dimensions[r].height = 18
            r += 1

    r += 1
    section_title(ws, r, 14, "B.  Day-of-Week Breakdown (all symbols, both windows combined)")
    r += 1
    dow_hdrs = ["Day", "N Signals", "Win Rate%", "Avg Outcome(pts)", "Expectancy(pts)",
                "Avg Win(pts)", "Avg Loss(pts)", "Verdict"]
    for ci, h in enumerate(dow_hdrs, 1):
        hdr(ws, r, ci, h)
    r += 1

    all_signals = pd.concat(
        [df for sym_d in gamma_data.values() for df in sym_d.values() if not df.empty],
        ignore_index=True
    )
    if not all_signals.empty and "day_of_week" in all_signals.columns:
        for dow in ["Monday", "Wednesday", "Friday"]:
            sub = all_signals[all_signals["day_of_week"] == dow]
            if sub.empty:
                continue
            s = summarise(sub)
            verdict = "BEST" if s["expectancy_pts"] > 50 else "GOOD" if s["expectancy_pts"] > 0 else "AVOID"
            vf = C_GREEN_MED if verdict == "BEST" else C_GREEN_LT if verdict == "GOOD" else C_RED_LT
            vals = [dow, s["count"], s["win_rate"], round(sub["outcome_pts"].mean(), 1),
                    s["expectancy_pts"], s["avg_win_pts"], s["avg_loss_pts"], verdict]
            for ci, v in enumerate(vals, 1):
                cell(ws, r, ci, v, fill=vf if ci == 8 else None)
            r += 1

    r += 1
    section_title(ws, r, 14, "C.  All Individual GammaSqueeze Signals")
    r += 1
    sig_hdrs = ["Symbol", "Window", "Date", "Day", "Time", "Direction",
                "Entry", "ADX14", "RSI14", "ATR", "Outcome(pts)", "WIN?"]
    for ci, h in enumerate(sig_hdrs, 1):
        hdr(ws, r, ci, h)
    r += 1

    for sym in SYMBOLS:
        for window in ("Train", "Test"):
            df = gamma_data.get(sym, {}).get(window, pd.DataFrame())
            if df.empty:
                continue
            for _, row in df.sort_values("date").iterrows():
                win = row.get("win", False)
                fg  = C_GREEN_LT if win else C_RED_LT
                vals = [
                    sym, window, row["date"],
                    row.get("day_of_week", "")[:3], row.get("time", ""),
                    row.get("direction", ""),
                    row.get("entry_price", ""), row.get("adx14", ""),
                    row.get("rsi14", ""),  row.get("atr", ""),
                    row.get("outcome_pts", ""),
                    "WIN" if win else "LOSS",
                ]
                for ci, v in enumerate(vals, 1):
                    cell(ws, r, ci, v,
                         fill=fg if ci == 12 else (C_GREEN_LT if win else C_RED_LT) if ci in (11,) else None,
                         halign="left" if ci in (1, 2, 3, 4, 5, 6) else "center")
                ws.row_dimensions[r].height = 16
                r += 1

    col_w(ws, [12, 8, 12, 10, 10, 10, 9, 8, 8, 8, 12, 8, 0, 0])
    freeze(ws, 3, 1)


# ── Sheet 5: GapFill Deep Analysis ────────────────────────────────────────────

def write_gap_fill_deep(ws, gap_data):
    """gap_data: {sym: {window: df_signals}}"""
    ws.title = "GapFill_Deep"
    ws.sheet_view.showGridLines = False

    title_row(ws, 1, 13,
              "GapFill Deep Analysis  |  BANKNIFTY + SENSEX  |  "
              "v1 Immediate-Open Entry  |  Min Gap 0.3%  |  Target: Prev Close  |  90-min Window",
              height=28)

    r = 3
    section_title(ws, r, 13, "A.  Overall Performance by Symbol and Window")
    r += 1
    for ci, h in enumerate(["Symbol", "Window", "N", "Fill Rate%", "Win Rate%",
                             "Avg Win(pts)", "Avg Loss(pts)", "Expectancy(pts)",
                             "UP Gap N", "UP WR%", "DN Gap N", "DN WR%", "Grade"], 1):
        hdr(ws, r, ci, h)
    r += 1

    for sym in ["BANKNIFTY", "SENSEX"]:
        for window in ("Train", "Test"):
            df = gap_data.get(sym, {}).get(window, pd.DataFrame())
            if df.empty:
                continue
            s    = summarise(df)
            fr   = df["filled"].mean() * 100 if "filled" in df.columns else 0
            up   = df[df["direction"] == "SHORT"]
            dn   = df[df["direction"] == "LONG"]
            grd, grd_fill = grade(s["win_rate"], s["expectancy_pts"], s["count"])
            vals = [sym, window, s["count"], round(fr, 1), s["win_rate"],
                    s["avg_win_pts"], s["avg_loss_pts"], s["expectancy_pts"],
                    len(up), round(up["win"].mean()*100, 1) if not up.empty else 0,
                    len(dn), round(dn["win"].mean()*100, 1) if not dn.empty else 0,
                    grd]
            for ci, v in enumerate(vals, 1):
                cell(ws, r, ci, v,
                     fill=(exp_fill(s["expectancy_pts"]) if ci == 8 else
                            wr_fill(s["win_rate"]) if ci == 5 else
                            grd_fill if ci == 13 else None))
            r += 1

    r += 1
    section_title(ws, r, 13, "B.  Gap Size Bucket Analysis")
    r += 1
    for ci, h in enumerate(["Symbol", "Window", "Gap Bucket", "N",
                             "Fill Rate%", "Win Rate%", "Expectancy(pts)",
                             "Avg Outcome(pts)", "Grade", "", "", "", ""], 1):
        hdr(ws, r, ci, h)
    r += 1

    for sym in ["BANKNIFTY", "SENSEX"]:
        for window in ("Train", "Test"):
            df = gap_data.get(sym, {}).get(window, pd.DataFrame())
            if df.empty or "gap_pct" not in df.columns:
                continue
            df = df.copy()
            df["bucket"] = pd.cut(df["gap_pct"].abs(),
                                   bins=[0, 0.5, 1.0, 999],
                                   labels=["0.3-0.5%", "0.5-1.0%", ">1.0%"])
            for bkt in ["0.3-0.5%", "0.5-1.0%", ">1.0%"]:
                sub = df[df["bucket"] == bkt]
                if sub.empty:
                    continue
                s   = summarise(sub)
                fr  = sub["filled"].mean() * 100 if "filled" in sub.columns else 0
                grd, grd_fill = grade(s["win_rate"], s["expectancy_pts"], s["count"])
                vals = [sym, window, bkt, s["count"], round(fr, 1), s["win_rate"],
                        s["expectancy_pts"], round(sub["outcome_pts"].mean(), 1), grd]
                for ci, v in enumerate(vals, 1):
                    cell(ws, r, ci, v,
                         fill=(grd_fill if ci == 9 else
                                exp_fill(s["expectancy_pts"]) if ci == 7 else None))
                r += 1

    r += 1
    section_title(ws, r, 13, "C.  Day-of-Week Win Rates (BANKNIFTY + SENSEX combined)")
    r += 1
    for ci, h in enumerate(["Day", "N", "Fill Rate%", "Win Rate%",
                             "Expectancy(pts)", "Avg Win", "Avg Loss", "Verdict",
                             "", "", "", "", ""], 1):
        hdr(ws, r, ci, h)
    r += 1

    all_gaps = pd.concat(
        [df for sym_d in gap_data.values() for df in sym_d.values() if not df.empty],
        ignore_index=True
    )
    if not all_gaps.empty and "date" in all_gaps.columns:
        all_gaps["dow"] = pd.to_datetime(all_gaps["date"]).dt.day_name().str[:3]
        for dow in ["Mon", "Tue", "Wed", "Thu", "Fri"]:
            sub = all_gaps[all_gaps["dow"] == dow]
            if sub.empty:
                continue
            s   = summarise(sub)
            fr  = sub["filled"].mean() * 100 if "filled" in sub.columns else 0
            vd  = "BEST" if s["expectancy_pts"] > 30 else "GOOD" if s["expectancy_pts"] > 0 else "WEAK"
            vf  = C_GREEN_MED if vd == "BEST" else C_GREEN_LT if vd == "GOOD" else C_RED_LT
            vals = [dow, s["count"], round(fr, 1), s["win_rate"],
                    s["expectancy_pts"], s["avg_win_pts"], s["avg_loss_pts"], vd]
            for ci, v in enumerate(vals, 1):
                cell(ws, r, ci, v, fill=vf if ci == 8 else None)
            r += 1

    r += 1
    section_title(ws, r, 13, "D.  All Individual GapFill Signals")
    r += 1
    for ci, h in enumerate(["Symbol", "Window", "Date", "Day", "Gap Dir",
                             "Gap%", "Prev Close", "Entry(Open)", "Filled?",
                             "Outcome(pts)", "WIN?", "Gap Bucket", ""], 1):
        hdr(ws, r, ci, h)
    r += 1

    for sym in ["BANKNIFTY", "SENSEX"]:
        for window in ("Train", "Test"):
            df = gap_data.get(sym, {}).get(window, pd.DataFrame())
            if df.empty:
                continue
            for _, row in df.sort_values("date").iterrows():
                win = row.get("win", False)
                fg  = C_GREEN_LT if win else C_RED_LT
                ap  = abs(row.get("gap_pct", 0))
                bkt = ("0.3-0.5%" if ap < 0.5 else "0.5-1.0%" if ap < 1 else ">1.0%")
                vals = [
                    sym, window, row["date"],
                    pd.to_datetime(row["date"]).strftime("%a"),
                    row.get("direction", ""),
                    round(row.get("gap_pct", 0), 3),
                    row.get("prev_close", ""),
                    row.get("entry_price", ""),
                    "YES" if row.get("filled", False) else "NO",
                    row.get("outcome_pts", ""),
                    "WIN" if win else "LOSS",
                    bkt, "",
                ]
                for ci, v in enumerate(vals, 1):
                    cell(ws, r, ci, v,
                         fill=(fg if ci in (9, 11) else None),
                         halign="left" if ci in (1, 2, 3, 4, 5) else "center")
                ws.row_dimensions[r].height = 16
                r += 1

    col_w(ws, [12, 8, 12, 5, 10, 8, 11, 12, 8, 12, 8, 10, 0])
    freeze(ws, 3, 1)


# ── Sheets 6-11: Daily Signal Sheets ─────────────────────────────────────────

def build_daily(df5m, symbol, v3_results):
    """Build one-row-per-day summary with v3 pattern signals."""
    daily = (df5m.groupby(df5m["timestamp"].dt.date)
             .agg(open=("open", "first"), high=("high", "max"),
                  low=("low", "min"),   close=("close", "last"))
             .reset_index().rename(columns={"timestamp": "date"}))
    daily["date_str"] = daily["date"].astype(str)
    daily["day"]      = pd.to_datetime(daily["date"]).dt.strftime("%a")
    daily["chg_pct"]  = round((daily["close"] - daily["open"]) / daily["open"] * 100, 2)

    def _sig(pat, date_str):
        df = v3_results.get(pat, pd.DataFrame())
        if df.empty:
            return "-"
        active = df[df.get("skipped", pd.Series([False]*len(df))).values == False] \
            if "skipped" in df.columns else df
        sub = active[active["date"] == date_str]
        if sub.empty:
            # check if skipped
            sk = df[df.get("skipped", pd.Series([False]*len(df))).values == True] \
                if "skipped" in df.columns else pd.DataFrame()
            if not sk.empty and (sk["date"] == date_str).any():
                return "SKIP"
            return "-"
        row = sub.iloc[0]
        cnt = len(sub)
        label = f"{row['direction'][:4]} {'WIN' if row['win'] else 'LOSS'}"
        return f"{label} (x{cnt})" if cnt > 1 else label

    rows = []
    EXP_WD = {"NIFTY": 1, "BANKNIFTY": 2, "SENSEX": 1}
    GAMMA_WD = {0, 2, 4}  # Mon, Wed, Fri
    exp_wd = EXP_WD.get(symbol, -1)

    for _, dr in daily.iterrows():
        ds = dr["date_str"]
        dt = pd.to_datetime(ds)
        is_exp   = dt.weekday() == exp_wd
        is_gamma = dt.weekday() in GAMMA_WD

        sigs = {p: _sig(p, ds) for p in V3_PATTERNS}
        wins   = sum(1 for v in sigs.values() if "WIN" in str(v))
        losses = sum(1 for v in sigs.values() if "LOSS" in str(v))
        skips  = sum(1 for v in sigs.values() if v == "SKIP")

        grade_str = ("GREAT"  if wins >= 2 and losses == 0 else
                     "WIN"    if wins > 0 and losses == 0   else
                     "LOSS"   if losses > 0 and wins == 0   else
                     "MIXED"  if wins > 0 and losses > 0    else
                     "EXPIRY" if is_exp                     else
                     "GAMMA"  if is_gamma                   else "-")

        rows.append({
            "date": ds, "day": dr["day"],
            "open": dr["open"], "high": dr["high"],
            "low": dr["low"],   "close": dr["close"], "chg": dr["chg_pct"],
            **sigs,
            "signals": wins + losses, "wins": wins, "losses": losses,
            "skips": skips, "grade": grade_str,
        })
    return pd.DataFrame(rows)


def write_daily_sheet(ws, df_daily, symbol, window_label):
    ws.title = f"{symbol}_{window_label}"
    ws.sheet_view.showGridLines = False

    title_row(ws, 1, 17,
              f"{symbol}  |  {window_label}  |  v3 Daily Signal Log  "
              f"| Green=WIN  Red=LOSS  Blue=Gamma/Expiry  Grey=SKIP", height=24)

    cols = ["Date", "Day", "Open", "High", "Low", "Close", "Chg%",
            "ExpiryBlast", "ORB15", "GapFill", "ATRSqueeze",
            "VWAPReclaim", "GammaSqueeze", "Sigs", "Wins", "Losses", "Grade"]
    for ci, h in enumerate(cols, 1):
        hdr(ws, 2, ci, h, rowh=26)

    EXP_WD    = {"NIFTY": 1, "BANKNIFTY": 2, "SENSEX": 1}
    GAMMA_WD  = {0, 2, 4}
    exp_wd    = EXP_WD.get(symbol, -1)

    for _, dr in df_daily.iterrows():
        r = ws.max_row + 1
        wins   = dr["wins"]
        losses = dr["losses"]
        grade_str = dr["grade"]
        is_exp   = pd.to_datetime(dr["date"]).weekday() == exp_wd
        is_gamma = pd.to_datetime(dr["date"]).weekday() in GAMMA_WD

        row_fill = (C_GREEN_LT   if wins > 0 and losses == 0   else
                    C_RED_LT     if losses > 0 and wins == 0   else
                    C_YELLOW     if wins > 0 and losses > 0    else
                    C_BLUE_LT    if is_gamma or is_exp         else None)

        grade_fill = (C_GREEN_MED  if grade_str == "GREAT"  else
                      C_GREEN_LT   if grade_str == "WIN"    else
                      C_RED_LT     if grade_str == "LOSS"   else
                      C_YELLOW     if grade_str == "MIXED"  else
                      C_BLUE_LT    if grade_str in ("EXPIRY", "GAMMA") else C_GREY_LT)

        row_vals = [
            dr["date"], dr["day"],
            dr["open"], dr["high"], dr["low"], dr["close"], dr["chg"],
            dr.get("ExpiryBlast_v2", "-"), dr.get("ORB15_v2", "-"),
            dr.get("GapFill_v2", "-"),     dr.get("ATRSqueeze_v2", "-"),
            dr.get("VWAPReclaim_v2", "-"), dr.get("GammaSqueeze", "-"),
            dr["signals"], dr["wins"], dr["losses"], grade_str,
        ]
        for ci, v in enumerate(row_vals, 1):
            if ci in range(8, 14):   # pattern columns
                if isinstance(v, str) and "WIN" in v:
                    f = C_GREEN_LT
                elif isinstance(v, str) and "LOSS" in v:
                    f = C_RED_LT
                elif isinstance(v, str) and v == "SKIP":
                    f = C_GREY_LT
                else:
                    f = row_fill
            elif ci == 17:
                f = grade_fill
            else:
                f = row_fill
            cell(ws, r, ci, v, fill=f,
                 halign="left"  if ci in (1, 2, 8, 9, 10, 11, 12, 13, 17) else "right")
        ws.row_dimensions[r].height = 16

    col_w(ws, [12, 4, 8, 8, 8, 8, 6, 14, 14, 14, 14, 14, 15, 5, 5, 6, 7])
    freeze(ws, 3, 1)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Strategy Assessment v3 -- Building Excel")
    print("=" * 60)

    # ── Load all data ─────────────────────────────────────────────────────────
    print("\n[1/6] Loading 5m data for all symbols ...")
    data = {}
    for sym in SYMBOLS:
        data[sym] = {}
        for label, f, t in [("Train", TRAIN_FROM, TRAIN_TO),
                             ("Test",  TEST_FROM,  TEST_TO)]:
            print(f"  {sym} {label} ({f} -> {t}) ...", end=" ", flush=True)
            df = load_and_compute(sym, f, t)
            data[sym][label] = df
            print(f"{len(df)} bars" if not df.empty else "no data")

    # ── Run patterns ──────────────────────────────────────────────────────────
    print("\n[2/6] Running v1 and v3 patterns ...")
    stats = {}
    gamma_data = {sym: {} for sym in SYMBOLS}
    gap_data   = {sym: {} for sym in ["BANKNIFTY", "SENSEX"]}
    v3_results = {sym: {} for sym in SYMBOLS}   # for daily sheets

    def _active(df):
        if df.empty or "skipped" not in df.columns:
            return df
        return df[df["skipped"] == False]

    for sym in SYMBOLS:
        stats[sym] = {}
        for label in ("Train", "Test"):
            df = data[sym][label]
            stats[sym][label] = {"v1": {}, "v3": {}}

            if df.empty:
                continue

            # v1
            v1 = run_v1(df, sym)
            for pat, dft in v1.items():
                stats[sym][label]["v1"][pat] = summarise(dft)

            # v3
            v3 = run_v3(df, sym)
            v3_results[sym][label] = v3
            for pat, dft in v3.items():
                active = _active(dft)
                s = summarise(active)
                stats[sym][label]["v3"][pat] = s
                if s["count"] > 0:
                    print(f"  {sym} {label} {pat}: {s['count']} signals  "
                          f"WR={s['win_rate']:.1f}%  Exp={s['expectancy_pts']:+.1f}pts")

            # cache for deep-dive sheets
            gs = _active(v3.get("GammaSqueeze", pd.DataFrame()))
            gamma_data[sym][label] = gs

            if sym in ("BANKNIFTY", "SENSEX"):
                gap_data[sym][label] = _active(v3.get("GapFill_v2", pd.DataFrame()))

    # ── Build Excel ───────────────────────────────────────────────────────────
    print("\n[3/6] Building Excel workbook ...")
    wb = Workbook()
    wb.remove(wb.active)

    print("  Writing Executive_Summary ...")
    write_executive_summary(wb.create_sheet("Executive_Summary"), stats)

    print("  Writing Final_Configs ...")
    write_final_configs(wb.create_sheet("Final_Configs"))

    print("  Writing Strategy_Evolution ...")
    write_evolution(wb.create_sheet("Strategy_Evolution"), stats)

    print("  Writing GammaSqueeze ...")
    write_gamma_squeeze(wb.create_sheet("GammaSqueeze"), gamma_data)

    print("  Writing GapFill_Deep ...")
    write_gap_fill_deep(wb.create_sheet("GapFill_Deep"), gap_data)

    # Daily signal sheets
    for sym in SYMBOLS:
        for label in ("6M", "7to12M"):
            window = "Train" if label == "6M" else "Test"
            df5m = data[sym][window]
            v3r  = v3_results[sym].get(window, {})
            if df5m.empty:
                continue
            print(f"  Writing {sym}_{label} ...")
            df_daily = build_daily(df5m, sym, v3r)
            write_daily_sheet(wb.create_sheet(f"{sym}_{label}"), df_daily, sym, label)

    # ── Save ──────────────────────────────────────────────────────────────────
    print(f"\n[4/6] Saving to:\n  {OUTPUT}")
    wb.save(str(OUTPUT))
    kb = OUTPUT.stat().st_size // 1024
    print(f"\nDone -- {kb} KB")
    print(f"\nOpen: {OUTPUT}")


if __name__ == "__main__":
    main()
