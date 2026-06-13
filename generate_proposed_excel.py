"""
generate_proposed_excel.py
--------------------------
Builds Strategy_Praposed_20260531_v101.xlsx with:
  1. Cover
  2. Summary_Comparison
  3. Improvements_Table
  4. GammaSqueeze
  5. NIFTY_6M
  6. NIFTY_7to12M
  7. BANKNIFTY_6M
  8. BANKNIFTY_7to12M
  9. SENSEX_6M
  10. SENSEX_7to12M

All text is ASCII-safe (no Unicode > 127).
Python: C:\ProgramData\anaconda3\python.exe
"""

import sys
import warnings
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_ROOT    = Path(__file__).parent
_MASTER  = _ROOT.parent / "MasterConfiguration"
_LIB     = _MASTER / "lib"
if str(_LIB) not in sys.path:
    sys.path.insert(0, str(_LIB))

OUTPUT_PATH = Path(
    r"C:\Users\meetm\OneDrive\Desktop\GCPPythonCode\MasterConfiguration"
    r"\reports\Strategy_Praposed_20260531_v101.xlsx"
)
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Imports from project
# ---------------------------------------------------------------------------

from historical_data_fetch import load
from pattern_discovery import add_indicators, summarise
from pattern_discovery_v2 import (
    add_indicators_v2,
    find_expiry_blast_v2,
    find_orb15_v2,
    find_gap_fill_v2,
    find_atr_squeeze_v2,
    find_vwap_reclaim_v2,
    find_gamma_squeeze,
    run_all_v2,
)
from pattern_discovery import (
    find_expiry_blast,
    find_orb15,
    find_gap_fill,
    find_atr_squeeze,
    find_vwap_reclaim,
)

# ---------------------------------------------------------------------------
# Date windows
# ---------------------------------------------------------------------------

TRAIN_FROM = "2025-11-26"
TRAIN_TO   = "2026-05-31"
TEST_FROM  = "2025-05-24"
TEST_TO    = "2025-11-26"

SYMBOLS = ["NIFTY", "BANKNIFTY", "SENSEX"]

# ---------------------------------------------------------------------------
# Color palette (hex strings without #)
# ---------------------------------------------------------------------------

C_HDR_BG     = "2E75B6"   # header row background
C_HDR_FG     = "FFFFFF"   # header font white
C_WIN        = "C6EFCE"   # WIN cell green
C_LOSS       = "FFC7CE"   # LOSS cell red
C_SKIP       = "F2F2F2"   # SKIP cell grey
C_BETTER     = "C6EFCE"   # proposed better than original
C_WORSE      = "FFC7CE"   # proposed worse
C_NEUTRAL    = "FFEB9C"   # neutral/same
C_TITLE_BG   = "1F4E79"   # title bar navy
C_SEC_BG     = "4472C4"   # section header blue
C_GAMMA_DAY  = "BDD7EE"   # gamma/expiry day light blue
C_WHITE      = "FFFFFF"
C_ROW_GREEN  = "E2EFDA"   # all-win row
C_ROW_RED    = "FFCCCC"   # all-loss row
C_ROW_YELLOW = "FFFF99"   # mixed row


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_row(ws, row_num: int, values: list, bg: str = C_HDR_BG):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill  = _fill(bg)
        cell.font  = _font(bold=True, color=C_HDR_FG)
        cell.alignment = _align("center")
        cell.border = _thin_border()


def _auto_width(ws, min_w: int = 8, max_w: int = 30):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value is not None:
                best = max(best, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = best


# ---------------------------------------------------------------------------
# Data loading helpers
# ---------------------------------------------------------------------------

def _load_and_prep(symbol: str, interval: str, from_date: str, to_date: str,
                   v2: bool = True) -> pd.DataFrame:
    df = load(symbol, interval, from_date, to_date)
    if df.empty:
        print(f"  [WARN] No data: {symbol} {interval} {from_date} -> {to_date}")
        return df
    df = df.sort_values("timestamp").reset_index(drop=True)
    if v2:
        df = add_indicators_v2(df)
    else:
        df = add_indicators(df)
    return df


def _daily_ohlc(symbol: str, from_date: str, to_date: str) -> pd.DataFrame:
    """Return daily OHLC from 5m data (first open, last close, day high/low)."""
    df5m = load(symbol, "5min", from_date, to_date)
    if df5m.empty:
        return pd.DataFrame()
    df5m["date"] = df5m["timestamp"].dt.date
    grp = df5m.groupby("date")
    daily = grp.agg(
        open  = ("open",  "first"),
        high  = ("high",  "max"),
        low   = ("low",   "min"),
        close = ("close", "last"),
        volume= ("volume","sum"),
    ).reset_index()
    daily["date"] = pd.to_datetime(daily["date"])
    daily["chg_pct"] = daily["close"].pct_change() * 100
    return daily


# ---------------------------------------------------------------------------
# Run all patterns for one symbol+window
# ---------------------------------------------------------------------------

def _run_patterns(symbol: str, from_date: str, to_date: str):
    """Returns dict: pattern_name -> DataFrame of signals (active only)."""
    df5m = _load_and_prep(symbol, "5min", from_date, to_date, v2=True)
    if df5m.empty:
        return {}, df5m

    results_v2 = run_all_v2(df5m, symbol)

    # Also run v1 for comparison
    df5m_v1 = df5m.copy()   # already has add_indicators applied inside add_indicators_v2
    results_v1 = {
        "ExpiryBlast": find_expiry_blast(df5m_v1, symbol),
        "ORB15":       find_orb15(df5m_v1, symbol),
        "GapFill":     find_gap_fill(df5m_v1, symbol),
        "ATRSqueeze":  find_atr_squeeze(df5m_v1, symbol),
        "VWAPReclaim": find_vwap_reclaim(df5m_v1, symbol),
    }

    return results_v2, results_v1, df5m


# ---------------------------------------------------------------------------
# Build daily signal map
# ---------------------------------------------------------------------------

PATTERN_COLS = [
    "ExpiryBlast_v2", "ORB15_v2", "GapFill_v2",
    "ATRSqueeze_v2", "VWAPReclaim_v2", "GammaSqueeze"
]


def _build_daily_signal_map(results_v2: dict) -> dict:
    """
    For each pattern, collapse multiple intraday signals to one per day.
    Returns: dict[date_str][pattern] = (label, win, skipped)
      label in {"BULL WIN","BULL LOSS","BEAR WIN","BEAR LOSS","SKIP","-"}
    """
    day_map = {}

    for pat_name in PATTERN_COLS:
        df_t = results_v2.get(pat_name, pd.DataFrame())
        if df_t.empty:
            continue

        for _, row in df_t.iterrows():
            date_str = str(row.get("date", ""))[:10]
            if date_str not in day_map:
                day_map[date_str] = {}

            # Skip rows
            if row.get("skipped", False):
                day_map[date_str][pat_name] = ("SKIP", False, True)
                continue

            direction = str(row.get("direction", "BULL"))
            bull = "BULL" if "BULL" in direction.upper() or direction.upper() in ("LONG",) else "BEAR"
            win  = bool(row.get("win", False))
            label = f"{bull} {'WIN' if win else 'LOSS'}"

            # Only overwrite if no signal yet for this pattern on this day
            if pat_name not in day_map[date_str]:
                day_map[date_str][pat_name] = (label, win, False)

    return day_map


# ---------------------------------------------------------------------------
# Write daily sheet (NIFTY_6M style)
# ---------------------------------------------------------------------------

def _write_daily_sheet(ws, symbol: str, from_date: str, to_date: str,
                       results_v2: dict, sheet_label: str):
    daily = _daily_ohlc(symbol, from_date, to_date)
    if daily.empty:
        ws.cell(1, 1).value = f"No data for {symbol} {from_date} -> {to_date}"
        return

    day_map = _build_daily_signal_map(results_v2)

    # Title row
    title_cell = ws.cell(1, 1, value=f"{sheet_label} -- {symbol}  [{from_date} to {to_date}]")
    title_cell.fill = _fill(C_TITLE_BG)
    title_cell.font = _font(bold=True, color=C_WHITE, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)

    # Header
    headers = [
        "Date", "Day", "Open", "High", "Low", "Close", "Chg%",
        "ExpiryBlast_v2", "ORB15_v2", "GapFill_v2",
        "ATRSqueeze_v2", "VWAPReclaim_v2", "GammaSqueeze",
        "Signals", "Wins", "Losses", "Grade"
    ]
    _hdr_row(ws, 2, headers)

    gamma_days = {0, 2, 4}   # Mon, Wed, Fri

    for r_idx, row in daily.iterrows():
        excel_row = r_idx + 3
        date_obj  = pd.Timestamp(row["date"])
        date_str  = date_obj.strftime("%Y-%m-%d")
        dow       = date_obj.strftime("%a")
        wd        = date_obj.weekday()

        pat_vals = {}
        for pat in PATTERN_COLS:
            info = day_map.get(date_str, {}).get(pat, ("-", False, False))
            pat_vals[pat] = info

        sig_count  = sum(1 for lbl, w, sk in pat_vals.values() if lbl not in ("-", "SKIP"))
        win_count  = sum(1 for lbl, w, sk in pat_vals.values() if w and not sk)
        loss_count = sum(1 for lbl, w, sk in pat_vals.values() if not w and lbl not in ("-", "SKIP") and not sk)

        grade = "-"
        if win_count > 0 and loss_count == 0:
            grade = "A"
        elif win_count > 0 and loss_count > 0:
            grade = "B"
        elif loss_count > 0 and win_count == 0:
            grade = "C"

        row_data = [
            date_str, dow,
            round(row["open"], 1), round(row["high"], 1),
            round(row["low"], 1),  round(row["close"], 1),
            round(row.get("chg_pct", 0) or 0, 2),
        ]
        for pat in PATTERN_COLS:
            lbl, _, _ = pat_vals[pat]
            row_data.append(lbl)
        row_data += [sig_count, win_count, loss_count, grade]

        # Row base fill
        if wd in gamma_days:
            row_fill = _fill(C_GAMMA_DAY)
        elif win_count > 0 and loss_count == 0:
            row_fill = _fill(C_ROW_GREEN)
        elif loss_count > 0 and win_count == 0:
            row_fill = _fill(C_ROW_RED)
        elif win_count > 0 and loss_count > 0:
            row_fill = _fill(C_ROW_YELLOW)
        else:
            row_fill = None

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(excel_row, col_idx, value=val)
            cell.border    = _thin_border()
            cell.alignment = _align("center")
            if row_fill:
                cell.fill = row_fill

        # Override pattern cells with WIN/LOSS/SKIP colors
        pat_col_start = 8   # column H = ExpiryBlast_v2
        for p_idx, pat in enumerate(PATTERN_COLS):
            lbl, win, skipped = pat_vals[pat]
            cell = ws.cell(excel_row, pat_col_start + p_idx)
            if lbl == "-":
                pass
            elif skipped or lbl == "SKIP":
                cell.fill = _fill(C_SKIP)
            elif win:
                cell.fill = _fill(C_WIN)
                cell.font = _font(bold=True)
            else:
                cell.fill = _fill(C_LOSS)
                cell.font = _font(bold=True)

    _auto_width(ws, min_w=10, max_w=22)
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20


# ---------------------------------------------------------------------------
# Sheet 1: Cover
# ---------------------------------------------------------------------------

def _write_cover(ws):
    ws.title = "Cover"

    data = [
        (1,  "STRATEGY ASSESSMENT -- PROPOSED v2 PATTERNS",  C_TITLE_BG, 16, True),
        (2,  "Generated: 2026-05-31",                         C_TITLE_BG, 12, False),
        (3,  "",                                               None,       11, False),
        (4,  "OVERVIEW",                                       C_SEC_BG,   13, True),
        (5,  "This report compares original (v1) vs proposed (v2) pattern parameters",  None, 11, False),
        (6,  "across three indices: NIFTY, BANKNIFTY, SENSEX.",                         None, 11, False),
        (7,  "Training window : 2025-11-26 to 2026-05-31 (6 months)",                  None, 11, False),
        (8,  "Test window     : 2025-05-24 to 2025-11-26 (months 7-12)",               None, 11, False),
        (9,  "",                                               None,       11, False),
        (10, "KEY IMPROVEMENTS SUMMARY",                       C_SEC_BG,   13, True),
        (11, "ExpiryBlast_v2  : Wider trigger window (14:45-15:15), BB pre-coiling filter, BANKNIFTY+SENSEX only", None, 11, False),
        (12, "ORB15_v2        : Gap filter <0.5%, tiered 3/6-bar hold, NIFTY+BANKNIFTY only", None, 11, False),
        (13, "GapFill_v2      : 5-min reversal candle entry, BANKNIFTY+SENSEX only",          None, 11, False),
        (14, "ATRSqueeze_v2   : Volume 1.8x SMA confirm, VIX-adaptive percentile",            None, 11, False),
        (15, "VWAPReclaim_v2  : Body% >0.65 filter, 3-bar (NIFTY) / 5-bar (BN,SX) confirm",  None, 11, False),
        (16, "GammaSqueeze    : NEW -- ADX>45, RSI 35-50, 3-bar surge, Mon/Wed/Fri only",     None, 11, False),
        (17, "",                                               None,       11, False),
        (18, "NAVIGATION",                                     C_SEC_BG,   13, True),
        (19, "Sheet 2  : Summary_Comparison    -- v1 vs v2 win-rate and expectancy", None, 11, False),
        (20, "Sheet 3  : Improvements_Table    -- parameter changes", None, 11, False),
        (21, "Sheet 4  : GammaSqueeze          -- new strategy results", None, 11, False),
        (22, "Sheet 5  : NIFTY_6M              -- NIFTY training signals (daily)", None, 11, False),
        (23, "Sheet 6  : NIFTY_7to12M          -- NIFTY test signals (daily)", None, 11, False),
        (24, "Sheet 7  : BANKNIFTY_6M          -- BANKNIFTY training", None, 11, False),
        (25, "Sheet 8  : BANKNIFTY_7to12M      -- BANKNIFTY test", None, 11, False),
        (26, "Sheet 9  : SENSEX_6M             -- SENSEX training", None, 11, False),
        (27, "Sheet 10 : SENSEX_7to12M         -- SENSEX test", None, 11, False),
        (28, "",                                               None,       11, False),
        (29, "COLOR LEGEND",                                   C_SEC_BG,   13, True),
    ]

    for r, text, bg, sz, bold in data:
        cell = ws.cell(r, 1, value=text)
        if bg:
            cell.fill = _fill(bg)
            cell.font = _font(bold=bold, color=C_WHITE, size=sz)
        else:
            cell.font = _font(bold=bold, size=sz)
        cell.alignment = _align("left")
    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")

    # Color legend swatches
    legend = [
        (30, C_WIN,       "WIN cell (light green)"),
        (31, C_LOSS,      "LOSS cell (light red)"),
        (32, C_SKIP,      "SKIP cell (grey -- new filter blocked signal)"),
        (33, C_GAMMA_DAY, "Gamma / expiry day row (light blue) -- Mon/Wed/Fri"),
        (34, C_ROW_GREEN, "Row highlight: all wins, no losses"),
        (35, C_ROW_RED,   "Row highlight: only losses"),
        (36, C_ROW_YELLOW,"Row highlight: mixed wins and losses"),
        (37, C_BETTER,    "Summary: proposed metric BETTER than original"),
        (38, C_WORSE,     "Summary: proposed metric WORSE than original"),
        (39, C_NEUTRAL,   "Summary: no change / neutral"),
    ]
    for r, color, desc in legend:
        swatch = ws.cell(r, 1)
        swatch.fill = _fill(color)
        swatch.border = _thin_border()
        swatch.value = "   "
        ws.cell(r, 2, value=desc).font = _font()

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 70


# ---------------------------------------------------------------------------
# Sheet 3: Improvements_Table
# ---------------------------------------------------------------------------

IMPROVEMENTS = [
    ("ExpiryBlast", "Trigger Window",      "15:00 IST (Rigid)",              "14:45-15:15 IST (Dynamic)",
     "Catches early/late expiry settlement volatility",          "BANKNIFTY, SENSEX"),
    ("ExpiryBlast", "Pre-Coiling Filter",  "None",                           "BB Width < 15th Percentile",
     "Restricts to highly compressed setups",                    "BANKNIFTY, SENSEX"),
    ("ExpiryBlast", "Target Indices",      "All",                            "SENSEX + BANKNIFTY Only",
     "Capitalizes on highest-expectancy indices",                "-"),
    ("ORB15",       "Target Indices",      "NIFTY, BANKNIFTY, SENSEX",       "NIFTY + BANKNIFTY Only",
     "Disables on SENSEX (no edge proven)",                      "-"),
    ("ORB15",       "Gap Filter",          "None",                           "Opening Gap < 0.5%",
     "Eliminates exhausted openings prone to whipsaws",          "NIFTY, BANKNIFTY"),
    ("ORB15",       "Failure Hold",        "30 minutes",                     "15 min conditional",
     "Cuts exposure on sluggish breakouts",                      "NIFTY, BANKNIFTY"),
    ("GapFill",     "Target Indices",      "NIFTY, BANKNIFTY, SENSEX",       "BANKNIFTY + SENSEX Only",
     "Restricts to high-fill-rate indices",                      "-"),
    ("GapFill",     "Entry Trigger",       "Immediate at open",              "5-min reversal candle close",
     "Prevents entering gaps that expand aggressively",          "BANKNIFTY, SENSEX"),
    ("ATRSqueeze",  "Squeeze Threshold",   "20th percentile (fixed)",        "VIX-Scaled 15th/25th percentile",
     "Auto-adjusts to market volatility regime",                 "All"),
    ("ATRSqueeze",  "Volume Confirm",      "None",                           "Volume > 1.8x 20-bar SMA",
     "Eliminates false breakouts lacking volume",                "All (bypassed for spot indices)"),
    ("VWAPReclaim", "Confirm Bars",        "3 consecutive",                  "3 (NIFTY) / 5 (BANKNIFTY, SENSEX)",
     "Higher barrier on volatile indices",                       "All"),
    ("VWAPReclaim", "Candle Body%",        "None",                           "Reclaim bar body > 65%",
     "Ensures institutional momentum at reclaim",                "All"),
    ("GammaSqueeze","NEW",                 "N/A",                            "ADX>45 + RSI 35-50 + 3-bar vol surge",
     "Captures gamma-hedging flow at expiry close",              "NIFTY, BANKNIFTY, SENSEX"),
]


def _write_improvements_table(ws):
    ws.title = "Improvements_Table"

    title = ws.cell(1, 1, value="Parameter Changes: Original (v1) vs Proposed (v2)")
    title.fill = _fill(C_TITLE_BG)
    title.font = _font(bold=True, color=C_WHITE, size=14)
    ws.merge_cells("A1:F1")

    headers = ["Pattern", "Parameter", "Original Value", "Proposed Value",
               "Expected Impact", "Applies To"]
    _hdr_row(ws, 2, headers)

    for r_idx, row in enumerate(IMPROVEMENTS, 3):
        pat = row[0]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, col_idx, value=val)
            cell.border    = _thin_border()
            cell.alignment = _align("left", wrap=True)
            if col_idx == 1:   # pattern name bold
                cell.font = _font(bold=True)
        # Highlight GammaSqueeze as new
        if pat == "GammaSqueeze":
            for c in range(1, 7):
                ws.cell(r_idx, c).fill = _fill(C_GAMMA_DAY)

    _auto_width(ws, min_w=12, max_w=45)
    ws.row_dimensions[1].height = 22


# ---------------------------------------------------------------------------
# Build comparison stats
# ---------------------------------------------------------------------------

def _compute_comparison():
    """
    Returns list of dicts with all columns needed for Summary_Comparison sheet.
    """
    rows = []
    patterns_v1 = ["ExpiryBlast", "ORB15", "GapFill", "ATRSqueeze", "VWAPReclaim"]
    patterns_v2 = ["ExpiryBlast_v2", "ORB15_v2", "GapFill_v2", "ATRSqueeze_v2", "VWAPReclaim_v2"]

    for sym in SYMBOLS:
        print(f"  Computing stats for {sym} ...")

        # Load and prep 5m data for both windows
        df_train_v2 = _load_and_prep(sym, "5min", TRAIN_FROM, TRAIN_TO, v2=True)
        df_test_v2  = _load_and_prep(sym, "5min", TEST_FROM,  TEST_TO,  v2=True)

        for p_v1, p_v2 in zip(patterns_v1, patterns_v2):

            # V1 train
            if not df_train_v2.empty:
                df_v1_train = _call_v1(p_v1, df_train_v2, sym)
                s_v1_train  = summarise(df_v1_train)
                df_v2_train = _call_v2(p_v2, df_train_v2, sym)
                active_v2_t = _active(df_v2_train)
                s_v2_train  = summarise(active_v2_t)
            else:
                s_v1_train = s_v2_train = {"win_rate": 0, "expectancy_pts": 0}

            # V1 test
            if not df_test_v2.empty:
                df_v1_test = _call_v1(p_v1, df_test_v2, sym)
                s_v1_test  = summarise(df_v1_test)
                df_v2_test = _call_v2(p_v2, df_test_v2, sym)
                active_v2_te = _active(df_v2_test)
                s_v2_test   = summarise(active_v2_te)
            else:
                s_v1_test = s_v2_test = {"win_rate": 0, "expectancy_pts": 0}

            rows.append({
                "Pattern":        p_v2,
                "Symbol":         sym,
                "Orig_Train_WR":  s_v1_train["win_rate"],
                "Prop_Train_WR":  s_v2_train["win_rate"],
                "WR_Delta":       round(s_v2_train["win_rate"] - s_v1_train["win_rate"], 1),
                "Orig_Train_Exp": s_v1_train["expectancy_pts"],
                "Prop_Train_Exp": s_v2_train["expectancy_pts"],
                "Exp_Delta":      round(s_v2_train["expectancy_pts"] - s_v1_train["expectancy_pts"], 1),
                "Orig_Test_WR":   s_v1_test["win_rate"],
                "Prop_Test_WR":   s_v2_test["win_rate"],
                "WR_Delta_Test":  round(s_v2_test["win_rate"] - s_v1_test["win_rate"], 1),
                "Orig_Test_Exp":  s_v1_test["expectancy_pts"],
                "Prop_Test_Exp":  s_v2_test["expectancy_pts"],
                "Exp_Delta_Test": round(s_v2_test["expectancy_pts"] - s_v1_test["expectancy_pts"], 1),
                "Net_Grade":      _net_grade(s_v2_train, s_v2_test),
            })

    return rows


def _active(df: pd.DataFrame) -> pd.DataFrame:
    """Filter out skipped signals."""
    if df.empty:
        return df
    if "skipped" in df.columns:
        return df[df["skipped"] == False].reset_index(drop=True)
    return df


def _call_v1(name: str, df5m: pd.DataFrame, sym: str) -> pd.DataFrame:
    fn_map = {
        "ExpiryBlast": find_expiry_blast,
        "ORB15":       find_orb15,
        "GapFill":     find_gap_fill,
        "ATRSqueeze":  find_atr_squeeze,
        "VWAPReclaim": find_vwap_reclaim,
    }
    try:
        return fn_map[name](df5m, sym)
    except Exception as e:
        print(f"    [WARN] v1 {name} {sym}: {e}")
        return pd.DataFrame()


def _call_v2(name: str, df5m: pd.DataFrame, sym: str) -> pd.DataFrame:
    fn_map = {
        "ExpiryBlast_v2": find_expiry_blast_v2,
        "ORB15_v2":       find_orb15_v2,
        "GapFill_v2":     find_gap_fill_v2,
        "ATRSqueeze_v2":  find_atr_squeeze_v2,
        "VWAPReclaim_v2": find_vwap_reclaim_v2,
        "GammaSqueeze":   find_gamma_squeeze,
    }
    try:
        return fn_map[name](df5m, sym)
    except Exception as e:
        print(f"    [WARN] v2 {name} {sym}: {e}")
        return pd.DataFrame()


def _net_grade(s_train: dict, s_test: dict) -> str:
    exp_t  = s_train.get("expectancy_pts", 0) or 0
    exp_te = s_test.get("expectancy_pts",  0) or 0
    wr_t   = s_train.get("win_rate",       0) or 0
    if exp_t > 10 and wr_t >= 55:
        return "STRONG"
    elif exp_t > 0 and wr_t >= 45:
        return "GOOD"
    elif exp_t > 0:
        return "MARGINAL"
    else:
        return "WEAK"


# ---------------------------------------------------------------------------
# Sheet 2: Summary_Comparison
# ---------------------------------------------------------------------------

def _write_summary_comparison(ws, comparison_rows: list):
    ws.title = "Summary_Comparison"

    title = ws.cell(1, 1, value="v1 vs v2 Pattern Comparison -- Win Rate and Expectancy")
    title.fill = _fill(C_TITLE_BG)
    title.font = _font(bold=True, color=C_WHITE, size=14)
    ws.merge_cells("A1:O1")

    headers = [
        "Pattern", "Symbol",
        "Orig_Train_WR", "Prop_Train_WR", "WR_Delta",
        "Orig_Train_Exp", "Prop_Train_Exp", "Exp_Delta",
        "Orig_Test_WR", "Prop_Test_WR", "WR_Delta_Test",
        "Orig_Test_Exp", "Prop_Test_Exp", "Exp_Delta_Test",
        "Net_Grade",
    ]
    _hdr_row(ws, 2, headers)

    grade_fill = {
        "STRONG":   _fill(C_WIN),
        "GOOD":     _fill(C_BETTER),
        "MARGINAL": _fill(C_NEUTRAL),
        "WEAK":     _fill(C_LOSS),
    }
    delta_cols = {5, 8, 11, 14}   # WR_Delta columns (1-indexed)

    for r_idx, row in enumerate(comparison_rows, 3):
        vals = [
            row["Pattern"], row["Symbol"],
            row["Orig_Train_WR"], row["Prop_Train_WR"], row["WR_Delta"],
            row["Orig_Train_Exp"], row["Prop_Train_Exp"], row["Exp_Delta"],
            row["Orig_Test_WR"], row["Prop_Test_WR"], row["WR_Delta_Test"],
            row["Orig_Test_Exp"], row["Prop_Test_Exp"], row["Exp_Delta_Test"],
            row["Net_Grade"],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(r_idx, col_idx, value=val)
            cell.border    = _thin_border()
            cell.alignment = _align("center")

            # Color delta cells
            if col_idx in delta_cols and isinstance(val, (int, float)):
                if val > 0:
                    cell.fill = _fill(C_BETTER)
                elif val < 0:
                    cell.fill = _fill(C_WORSE)
                else:
                    cell.fill = _fill(C_NEUTRAL)

        # Grade cell
        grade_cell = ws.cell(r_idx, 15)
        gf = grade_fill.get(row["Net_Grade"])
        if gf:
            grade_cell.fill = gf
        grade_cell.font = _font(bold=True)

    _auto_width(ws, min_w=10, max_w=18)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Sheet 4: GammaSqueeze results
# ---------------------------------------------------------------------------

def _write_gamma_squeeze_sheet(ws, gamma_results: dict):
    ws.title = "GammaSqueeze"

    title = ws.cell(1, 1, value="GammaSqueeze -- NEW Strategy  (Mon/Wed/Fri  14:45-15:15 IST)")
    title.fill = _fill(C_TITLE_BG)
    title.font = _font(bold=True, color=C_WHITE, size=14)
    ws.merge_cells("A1:K1")

    # ---- Stats summary table ----
    ws.cell(3, 1, value="STATS SUMMARY").fill = _fill(C_SEC_BG)
    ws.cell(3, 1).font = _font(bold=True, color=C_WHITE)
    ws.merge_cells("A3:K3")

    stat_hdrs = ["Symbol", "Window", "Signals", "Wins", "Losses",
                 "WinRate%", "AvgWin", "AvgLoss", "Expectancy", "Best", "Worst"]
    _hdr_row(ws, 4, stat_hdrs)

    stat_row = 5
    for sym in SYMBOLS:
        for wlabel, wfrom, wto in [("Train (6M)", TRAIN_FROM, TRAIN_TO),
                                    ("Test (7-12M)", TEST_FROM, TEST_TO)]:
            key = f"{sym}_{wlabel}"
            df_t = gamma_results.get(key, pd.DataFrame())
            if df_t.empty or "skipped" not in df_t.columns:
                active = df_t
            else:
                active = df_t[df_t["skipped"] == False]
            s = summarise(active)
            vals = [sym, wlabel, s["count"],
                    s.get("win_count", 0), s.get("loss_count", 0),
                    s["win_rate"],
                    s["avg_win_pts"], s["avg_loss_pts"],
                    s["expectancy_pts"],
                    s.get("best_pts", 0), s.get("worst_pts", 0)]
            for col_idx, v in enumerate(vals, 1):
                cell = ws.cell(stat_row, col_idx, value=v)
                cell.border    = _thin_border()
                cell.alignment = _align("center")
            # Color win rate
            wr_cell = ws.cell(stat_row, 6)
            if s["win_rate"] >= 55:
                wr_cell.fill = _fill(C_WIN)
            elif s["win_rate"] >= 45:
                wr_cell.fill = _fill(C_NEUTRAL)
            else:
                wr_cell.fill = _fill(C_LOSS)
            stat_row += 1

    # ---- Day-of-week breakdown ----
    dow_row = stat_row + 2
    ws.cell(dow_row, 1, value="DAY-OF-WEEK BREAKDOWN").fill = _fill(C_SEC_BG)
    ws.cell(dow_row, 1).font = _font(bold=True, color=C_WHITE)
    ws.merge_cells(f"A{dow_row}:K{dow_row}")
    dow_row += 1

    dow_hdrs = ["Symbol", "Window", "Day", "Signals", "Wins", "WinRate%", "Avg Outcome"]
    _hdr_row(ws, dow_row, dow_hdrs)
    dow_row += 1

    for sym in SYMBOLS:
        for wlabel, wfrom, wto in [("Train (6M)", TRAIN_FROM, TRAIN_TO),
                                    ("Test (7-12M)", TEST_FROM, TEST_TO)]:
            key = f"{sym}_{wlabel}"
            df_t = gamma_results.get(key, pd.DataFrame())
            if df_t.empty:
                continue
            active = df_t[df_t.get("skipped", pd.Series([False]*len(df_t))) == False] if "skipped" in df_t.columns else df_t
            if active.empty:
                continue
            for day_name in ["Monday", "Wednesday", "Friday"]:
                sub = active[active["day_of_week"] == day_name] if "day_of_week" in active.columns else pd.DataFrame()
                n   = len(sub)
                if n == 0:
                    continue
                wins = sub["win"].sum() if "win" in sub.columns else 0
                wr   = round(wins / n * 100, 1)
                avg  = round(sub["outcome_pts"].mean(), 1) if "outcome_pts" in sub.columns else 0
                for col_idx, v in enumerate([sym, wlabel, day_name, n, int(wins), wr, avg], 1):
                    cell = ws.cell(dow_row, col_idx, value=v)
                    cell.border    = _thin_border()
                    cell.alignment = _align("center")
                dow_row += 1

    # ---- Individual signals ----
    sig_row = dow_row + 2
    ws.cell(sig_row, 1, value="INDIVIDUAL SIGNALS -- TRAINING WINDOW").fill = _fill(C_SEC_BG)
    ws.cell(sig_row, 1).font = _font(bold=True, color=C_WHITE)
    ws.merge_cells(f"A{sig_row}:K{sig_row}")
    sig_row += 1

    sig_hdrs = ["Symbol", "Date", "Day", "Time", "Direction",
                "Entry", "ADX14", "RSI14", "ATR", "Outcome_Pts", "Win"]
    _hdr_row(ws, sig_row, sig_hdrs)
    sig_row += 1

    for sym in SYMBOLS:
        key = f"{sym}_Train (6M)"
        df_t = gamma_results.get(key, pd.DataFrame())
        if df_t.empty:
            continue
        active = df_t[df_t["skipped"] == False] if "skipped" in df_t.columns else df_t
        for _, row in active.iterrows():
            vals = [
                sym,
                str(row.get("date", ""))[:10],
                row.get("day_of_week", ""),
                str(row.get("time", "")),
                row.get("direction", ""),
                row.get("entry_price", ""),
                row.get("adx14", ""),
                row.get("rsi14", ""),
                row.get("atr", ""),
                row.get("outcome_pts", ""),
                "WIN" if row.get("win") else "LOSS",
            ]
            for col_idx, v in enumerate(vals, 1):
                cell = ws.cell(sig_row, col_idx, value=v)
                cell.border    = _thin_border()
                cell.alignment = _align("center")
            win_cell = ws.cell(sig_row, 11)
            win_cell.fill = _fill(C_WIN if row.get("win") else C_LOSS)
            win_cell.font = _font(bold=True)
            sig_row += 1

    _auto_width(ws, min_w=8, max_w=20)
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

def main():
    print("=" * 62)
    print("  Strategy_Praposed_20260531_v101.xlsx  -- Building ...")
    print("=" * 62)

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ---- Pre-load all data ----
    print("\n[1/5] Loading 5m data for all symbols ...")

    all_data = {}   # sym -> {window -> (df5m_v2, results_v2, results_v1)}

    for sym in SYMBOLS:
        all_data[sym] = {}
        for wlabel, wfrom, wto in [("train", TRAIN_FROM, TRAIN_TO),
                                    ("test",  TEST_FROM,  TEST_TO)]:
            print(f"  {sym} {wlabel} ({wfrom} -> {wto}) ...")
            df5m = _load_and_prep(sym, "5min", wfrom, wto, v2=True)
            if df5m.empty:
                all_data[sym][wlabel] = (pd.DataFrame(), {}, {})
                continue
            results_v2 = run_all_v2(df5m, sym)
            results_v1 = {
                "ExpiryBlast": find_expiry_blast(df5m, sym),
                "ORB15":       find_orb15(df5m, sym),
                "GapFill":     find_gap_fill(df5m, sym),
                "ATRSqueeze":  find_atr_squeeze(df5m, sym),
                "VWAPReclaim": find_vwap_reclaim(df5m, sym),
            }
            all_data[sym][wlabel] = (df5m, results_v2, results_v1)
            # quick summary
            for pat, df_t in results_v2.items():
                if df_t.empty:
                    continue
                active = df_t[df_t["skipped"] == False] if "skipped" in df_t.columns else df_t
                s = summarise(active)
                print(f"    {pat:20s}: {s['count']:3d} signals  WR={s['win_rate']:5.1f}%  Exp={s['expectancy_pts']:+6.1f}pts")

    # ---- Sheet 1: Cover ----
    print("\n[2/5] Writing Cover sheet ...")
    ws_cover = wb.create_sheet("Cover")
    _write_cover(ws_cover)

    # ---- Sheet 2: Summary_Comparison ----
    print("[3/5] Computing comparison stats ...")
    comparison_rows = []
    patterns_v1 = ["ExpiryBlast", "ORB15", "GapFill", "ATRSqueeze", "VWAPReclaim"]
    patterns_v2 = ["ExpiryBlast_v2", "ORB15_v2", "GapFill_v2", "ATRSqueeze_v2", "VWAPReclaim_v2"]

    for sym in SYMBOLS:
        df_train_v2, rv2_train, rv1_train = all_data[sym].get("train", (pd.DataFrame(), {}, {}))
        df_test_v2,  rv2_test,  rv1_test  = all_data[sym].get("test",  (pd.DataFrame(), {}, {}))

        for p_v1, p_v2 in zip(patterns_v1, patterns_v2):
            def _s(d): return summarise(d) if not d.empty else {"win_rate": 0, "expectancy_pts": 0}
            def _a(d): return d[d["skipped"] == False] if not d.empty and "skipped" in d.columns else d

            s_v1_train = _s(rv1_train.get(p_v1, pd.DataFrame()))
            s_v2_train = _s(_a(rv2_train.get(p_v2, pd.DataFrame())))
            s_v1_test  = _s(rv1_test.get(p_v1,  pd.DataFrame()))
            s_v2_test  = _s(_a(rv2_test.get(p_v2,  pd.DataFrame())))

            comparison_rows.append({
                "Pattern":        p_v2,
                "Symbol":         sym,
                "Orig_Train_WR":  s_v1_train["win_rate"],
                "Prop_Train_WR":  s_v2_train["win_rate"],
                "WR_Delta":       round(s_v2_train["win_rate"] - s_v1_train["win_rate"], 1),
                "Orig_Train_Exp": s_v1_train["expectancy_pts"],
                "Prop_Train_Exp": s_v2_train["expectancy_pts"],
                "Exp_Delta":      round(s_v2_train["expectancy_pts"] - s_v1_train["expectancy_pts"], 1),
                "Orig_Test_WR":   s_v1_test["win_rate"],
                "Prop_Test_WR":   s_v2_test["win_rate"],
                "WR_Delta_Test":  round(s_v2_test["win_rate"] - s_v1_test["win_rate"], 1),
                "Orig_Test_Exp":  s_v1_test["expectancy_pts"],
                "Prop_Test_Exp":  s_v2_test["expectancy_pts"],
                "Exp_Delta_Test": round(s_v2_test["expectancy_pts"] - s_v1_test["expectancy_pts"], 1),
                "Net_Grade":      _net_grade(s_v2_train, s_v2_test),
            })

    ws_summary = wb.create_sheet("Summary_Comparison")
    _write_summary_comparison(ws_summary, comparison_rows)

    # ---- Sheet 3: Improvements_Table ----
    ws_imp = wb.create_sheet("Improvements_Table")
    _write_improvements_table(ws_imp)

    # ---- Sheet 4: GammaSqueeze ----
    print("[4/5] Writing GammaSqueeze sheet ...")
    gamma_results = {}
    for sym in SYMBOLS:
        for wlabel, wfrom, wto in [("Train (6M)", TRAIN_FROM, TRAIN_TO),
                                    ("Test (7-12M)", TEST_FROM, TEST_TO)]:
            w_key  = "train" if "Train" in wlabel else "test"
            df5m_v2, rv2, _ = all_data[sym].get(w_key, (pd.DataFrame(), {}, {}))
            gamma_results[f"{sym}_{wlabel}"] = rv2.get("GammaSqueeze", pd.DataFrame())

    ws_gamma = wb.create_sheet("GammaSqueeze")
    _write_gamma_squeeze_sheet(ws_gamma, gamma_results)

    # ---- Sheets 5-10: Daily signal sheets ----
    print("[5/5] Writing daily signal sheets ...")
    sheet_specs = [
        ("NIFTY_6M",         "NIFTY",     "train", TRAIN_FROM, TRAIN_TO),
        ("NIFTY_7to12M",     "NIFTY",     "test",  TEST_FROM,  TEST_TO),
        ("BANKNIFTY_6M",     "BANKNIFTY", "train", TRAIN_FROM, TRAIN_TO),
        ("BANKNIFTY_7to12M", "BANKNIFTY", "test",  TEST_FROM,  TEST_TO),
        ("SENSEX_6M",        "SENSEX",    "train", TRAIN_FROM, TRAIN_TO),
        ("SENSEX_7to12M",    "SENSEX",    "test",  TEST_FROM,  TEST_TO),
    ]

    for sheet_name, sym, wlabel, wfrom, wto in sheet_specs:
        print(f"  Writing {sheet_name} ...")
        ws = wb.create_sheet(sheet_name)
        _, rv2, _ = all_data[sym].get(wlabel, (pd.DataFrame(), {}, {}))
        _write_daily_sheet(ws, sym, wfrom, wto, rv2, sheet_name)

    # ---- Save ----
    print(f"\nSaving to:\n  {OUTPUT_PATH}")
    wb.save(str(OUTPUT_PATH))
    size_kb = OUTPUT_PATH.stat().st_size // 1024
    print(f"Saved. File size: {size_kb} KB")
    print(f"\nDone.  Open: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
